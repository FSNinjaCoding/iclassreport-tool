"""
iClassReport Processor  -  Flip Side Lowell, LLC
================================================
Upload two iClassPro exports and get the canonical iClassReport ST Ledgerwood
requires, a draft 2090 clearing JE, and a merchant-statement validation.

Inputs (CSV exports, run for the SETTLEMENT WINDOW shown in Step 1):
  A) FIN-4  Program Deposit Split Report
  B) FIN-24 Gateway Transactions Report

v1.3 changes
  - "Payout date" basis mode (recommended): export FIN-4 with iClassPro's
    Payout Date option for the plain calendar month - it natively ties to the
    merchant statement, no settlement-window math needed. Fees are keyed in
    from the statement (the gateway report cannot filter by payout on this
    account - Batch Date is unpopulated).
  - Accepts .xlsx exports as well as .csv (requires openpyxl in
    requirements.txt)

v1.2 changes
  - Step 1 banner computes the settlement window (statement no longer needed
    to know the export dates; it becomes the validation target instead)
  - Exact-match validation against the merchant statement (replaces the old
    "small variance expected" bank cross-check)
  - Draft JE is tender-filtered: built from gateway-settled tenders only
    (Credit Card / Swipe / Card Present / eCheck). Cash, Check, External CC,
    and Nacha are excluded and reported separately so nothing double-counts
    against direct bank-feed coding (e.g. ClassWallet) or next month's cash
    deposit.
  - JE now shows the DR 2090 line and self-checks that debits == credits
  - Warns if the uploaded reports look like a calendar month instead of the
    settlement window

Logic validated against June 2026 (ties to the merchant statement, bank, and
Xero to the penny on settlement basis).  Version 1.2
"""
import csv
import io
import re
import datetime

import pandas as pd
import streamlit as st

# --- CONFIGURATION ---------------------------------------------------------
VERSION = "1.3.1"

# Open Gym and Birthdays lead the report; every other taxable program follows
# (alphabetical), one row each. Taxable vs non-taxable is detected from the
# report (tax collected > 0), so new programs classify themselves.
OPEN_GYM_NAMES = {"open gym"}
BIRTHDAY_NAMES = {"birthdays", "birthday"}

# FIN-4 tender columns that settle through the iClassPro gateway (these are
# the dollars that arrive as "iClassPro Inc. PAYOUT" in the bank / 2090).
GATEWAY_TENDERS = ["Credit Card", "Credit Card - Swipe/insert/tap",
                   "Credit Card Present", "eCheck"]
CASH_TENDERS = ["Cash", "Check"]
EXTERNAL_TENDERS = ["External Credit Card", "Nacha"]

st.set_page_config(page_title=f"iClassReport Processor {VERSION}", page_icon="N", layout="wide")


# =========================================================================== #
#  SETTLEMENT WINDOW  (pure Python - no pandas holiday deps)
# =========================================================================== #
def _nth_weekday(year, month, weekday, n):
    d = datetime.date(year, month, 1)
    return d + datetime.timedelta(days=(weekday - d.weekday()) % 7 + 7 * (n - 1))


def _last_weekday(year, month, weekday):
    d = (datetime.date(year, 12, 31) if month == 12
         else datetime.date(year, month + 1, 1) - datetime.timedelta(days=1))
    return d - datetime.timedelta(days=(d.weekday() - weekday) % 7)


def _observed(d):
    """Federal observance: Saturday holidays observed Friday, Sunday ones Monday."""
    if d.weekday() == 5:
        return d - datetime.timedelta(days=1)
    if d.weekday() == 6:
        return d + datetime.timedelta(days=1)
    return d


def us_federal_holidays(year):
    hols = {
        _observed(datetime.date(year, 1, 1)),        # New Year's Day
        _nth_weekday(year, 1, 0, 3),                 # MLK Day
        _nth_weekday(year, 2, 0, 3),                 # Presidents Day
        _last_weekday(year, 5, 0),                   # Memorial Day
        _observed(datetime.date(year, 6, 19)),       # Juneteenth
        _observed(datetime.date(year, 7, 4)),        # Independence Day
        _nth_weekday(year, 9, 0, 1),                 # Labor Day
        _nth_weekday(year, 10, 0, 2),                # Columbus Day
        _observed(datetime.date(year, 11, 11)),      # Veterans Day
        _nth_weekday(year, 11, 3, 4),                # Thanksgiving
        _observed(datetime.date(year, 12, 25)),      # Christmas
    }
    hols.add(_observed(datetime.date(year + 1, 1, 1)))   # next New Year's can observe Dec 31
    return hols


def _is_bd(d, hols):
    return d.weekday() < 5 and d not in hols


def _roll_fwd(d, hols):
    while not _is_bd(d, hols):
        d += datetime.timedelta(days=1)
    return d


def _minus_bd(d, n, hols):
    while n > 0:
        d -= datetime.timedelta(days=1)
        if _is_bd(d, hols):
            n -= 1
    return d


def settlement_window(year: int, month: int):
    """iClassPro pays out ~2 business days after the transaction, so the
    money that lands in the bank during a month comes from this window.
    Windows tile perfectly: no gaps, no overlaps."""
    hols = us_federal_holidays(year - 1) | us_federal_holidays(year) | us_federal_holidays(year + 1)
    first_bd = _roll_fwd(datetime.date(year, month, 1), hols)
    start = _minus_bd(first_bd, 2, hols)
    ny, nm = (year + 1, 1) if month == 12 else (year, month + 1)
    next_first_bd = _roll_fwd(datetime.date(ny, nm, 1), hols)
    end = _minus_bd(next_first_bd, 2, hols) - datetime.timedelta(days=1)
    return start, end


# =========================================================================== #
#  CORE LOGIC  (pure functions)
# =========================================================================== #
def parse_money(s):
    """'$1,234.56' / '$(98.55)' / '--' / '' -> float."""
    s = str(s).replace("\xa0", " ").replace("$", "").replace(",", "").strip()
    if s in ("", "--"):
        return 0.0
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    try:
        v = float(s)
    except ValueError:
        return 0.0
    return -v if neg else v


def _read_rows(raw_text):
    raw_text = raw_text.lstrip("\ufeff")
    return list(csv.reader(io.StringIO(raw_text)))


def load_upload(uploaded_file):
    """Streamlit UploadedFile (.csv or .xlsx) -> list-of-rows of strings."""
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)
    if name.endswith((".xlsx", ".xls")):
        try:
            from openpyxl import load_workbook
        except ImportError:
            st.error("Excel uploads need `openpyxl` - add a line saying `openpyxl` "
                     "to requirements.txt in GitHub, or export as CSV instead.")
            return None
        wb = load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
        rows = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            sheet_rows = [["" if c is None else str(c) for c in r]
                          for r in ws.iter_rows(values_only=True)]
            if any(any(c.strip() for c in r) for r in sheet_rows):
                rows = sheet_rows
                break
        if not rows:
            st.error(f"'{uploaded_file.name}' contains no data. If this was a "
                     "payout-filtered Gateway export, that's expected - the gateway "
                     "report can't filter by payout date (Batch Date is unpopulated). "
                     "Use statement fees instead.")
            return None
        return rows
    return _read_rows(uploaded_file.read().decode("utf-8-sig", errors="ignore"))


def detect_report_type(rows):
    """Return 'gateway', 'program_split', or None by inspecting the first rows."""
    flat = " ".join(" ".join(str(c) for c in r) for r in rows[:8]).lower()
    if "final settlement amount" in flat or "transaction id" in flat:
        return "gateway"
    if "charge category" in flat or "payments received" in flat:
        return "program_split"
    return None


def parse_program_split(rows):
    """FIN-4 -> (dict program -> {total, tax, gross, refunds[], gw, cash, external},
                 unapplied_gw_net, period_str)

    gw / cash / external are the Program Total row's tender-column dollars:
      gw       = tenders that settle through the iClassPro gateway
      cash     = Cash + Check (deposited manually, hits the bank next month)
      external = External Credit Card + Nacha (settled outside the gateway)
    unapplied_gw_net = net gateway-tender dollars on the Unapplied rows.
    """
    period = ""
    for row in rows[:6]:
        for cell in row:
            m = re.search(r"\d{2}/\d{2}/\d{4}\s*-\s*\d{2}/\d{2}/\d{4}", str(cell))
            if m:
                period = m.group(0)
                break
        if period:
            break

    hidx = None
    for i, r in enumerate(rows[:10]):
        low = [str(c).strip().lower() for c in r]
        if "program" in low and "charge category" in low:
            hidx = i
            break
    if hidx is None:
        hidx = 1
    header = [c.strip() for c in rows[hidx]]

    def col(name, default=None):
        for i, c in enumerate(header):
            if c.strip().lower() == name.lower():
                return i
        return default

    i_prog = col("Program", 1)
    i_cat = col("Charge Category", 2)
    i_total = col("Total", 3)
    i_tax = col("Taxes", 4)
    i_gw = [col(n) for n in GATEWAY_TENDERS if col(n) is not None]
    i_cash = [col(n) for n in CASH_TENDERS if col(n) is not None]
    i_ext = [col(n) for n in EXTERNAL_TENDERS if col(n) is not None]

    def tender_sum(row, idxs):
        return sum(parse_money(row[i]) for i in idxs if len(row) > i)

    progs = {}
    unapplied_gw_net = 0.0
    cur = None
    for r in rows[hidx + 1:]:
        if len(r) <= i_cat:
            continue
        name = r[i_prog].strip()
        cat = r[i_cat].strip()
        if name.lower().startswith("total payments received"):
            break
        if name.lower().startswith("unapplied"):
            unapplied_gw_net += tender_sum(r, i_gw)
            continue
        if cat == "Program Total:":
            cur = name
            progs[cur] = {
                "total": parse_money(r[i_total]),
                "tax": parse_money(r[i_tax]),
                "gw": tender_sum(r, i_gw),
                "cash": tender_sum(r, i_cash),
                "external": tender_sum(r, i_ext),
                "gross": 0.0,
                "refunds": [],
            }
        elif cur and name == "" and cat:
            amt = parse_money(r[i_total])
            if "refund" in cat.lower():
                progs[cur]["refunds"].append(amt)
            else:
                progs[cur]["gross"] += amt
    return progs, round(unapplied_gw_net, 2), period


def parse_gateway(rows):
    """FIN-24 -> dict with gross/fees/net + settled/unsettled splits + period."""
    hidx = None
    for i, r in enumerate(rows[:10]):
        low = [str(c).strip().lower() for c in r]
        if "amount" in low and "final settlement amount" in low:
            hidx = i
            break
    if hidx is None:
        hidx = 0
    header = [c.strip() for c in rows[hidx]]

    def col(name):
        for i, c in enumerate(header):
            if c.strip().lower() == name.lower():
                return i
        return None

    i_date = col("Date")
    i_amt = col("Amount")
    i_fee = col("Fees")
    i_net = col("Final Settlement Amount")
    i_status = col("Status")

    gross = fees = net = settled_net = unsettled_amt = 0.0
    n = unsettled_n = 0
    dates = []
    for r in rows[hidx + 1:]:
        if i_amt is None or len(r) <= i_amt:
            continue
        if i_date is not None and (len(r) <= i_date or "/" not in str(r[i_date])):
            continue  # skip totals row
        a = parse_money(r[i_amt])
        f = parse_money(r[i_fee]) if i_fee is not None else 0.0
        s = parse_money(r[i_net]) if i_net is not None else (a - f)
        status = (r[i_status].strip().lower() if i_status is not None and len(r) > i_status else "")
        gross += a
        fees += f
        net += s
        n += 1
        if i_date is not None:
            dates.append(str(r[i_date]))
        if status == "settled":
            settled_net += s
        else:
            unsettled_amt += a
            unsettled_n += 1
    period = ""
    dmin = dmax = None
    if dates:
        try:
            ds = [datetime.datetime.strptime(d, "%m/%d/%Y") for d in dates]
            dmin, dmax = min(ds).date(), max(ds).date()
            period = f"{dmin:%m/%d/%Y} - {dmax:%m/%d/%Y}"
        except Exception:
            period = ""
    return {
        "gross": round(gross, 2),
        "fees": round(fees, 2),
        "net": round(net, 2),
        "count": n,
        "settled_net": round(settled_net, 2),
        "unsettled_amt": round(unsettled_amt, 2),
        "unsettled_count": unsettled_n,
        "period": period,
        "date_min": dmin,
        "date_max": dmax,
    }


def build_iclassreport(progs, unapplied_gw_net, gateway, use_tax=0.0):
    taxable = [p for p, v in progs.items() if abs(v["tax"]) > 0.001]
    open_gym = [p for p in progs if p.strip().lower() in OPEN_GYM_NAMES]
    birthdays = [p for p in progs if p.strip().lower() in BIRTHDAY_NAMES]

    ref_taxable = sum(-x for p in taxable for x in progs[p]["refunds"])
    ref_nontaxable = sum(-x for p in progs if p not in taxable for x in progs[p]["refunds"])

    # one row per taxable program: Open Gym, Birthdays, then the rest alphabetically
    ordered = list(open_gym) + list(birthdays) + sorted(
        p for p in taxable if p not in open_gym and p not in birthdays)
    program_rows = [(p, round(progs[p]["gross"], 2), round(progs[p]["tax"], 2)) for p in ordered]
    total_tax = round(sum(progs[p]["tax"] for p in taxable), 2)

    # ---- Draft 2090 clearing JE - GATEWAY-SETTLED TENDERS ONLY -------------
    # Each program's tax is allocated to the gateway share pro-rata (most
    # programs are 100% card, so this only matters for mixed cash/card ones).
    retail_progs = [p for p in taxable if p not in open_gym and p not in birthdays]

    def gw_tax(p):
        v = progs[p]
        return v["tax"] * (v["gw"] / v["total"]) if abs(v["total"]) > 0.001 else 0.0

    venue = round(sum(progs[p]["gw"] - gw_tax(p) for p in open_gym + birthdays), 2)
    merch = round(sum(progs[p]["gw"] - gw_tax(p) for p in retail_progs), 2)
    member = round(sum(progs[p]["gw"] for p in progs
                       if p not in taxable) + unapplied_gw_net, 2)
    je_tax = round(sum(gw_tax(p) for p in progs), 2)

    je_credits = round(venue + merch + member + je_tax, 2)
    je_debits = round(gateway["net"] + gateway["fees"], 2)
    # absorb sub-dollar pro-rata rounding into the largest revenue line
    rounding = round(je_debits - je_credits, 2)
    if abs(rounding) <= 1.00 and abs(rounding) > 0.001:
        member = round(member + rounding, 2)
        je_credits = round(venue + merch + member + je_tax, 2)

    cash_collected = round(sum(v["cash"] for v in progs.values()), 2)
    external_total = round(sum(v["external"] for v in progs.values()), 2)

    notes = []
    extra_taxable = [p for p in taxable if p not in open_gym and p not in birthdays]
    if extra_taxable:
        notes.append("Taxable programs beyond Open Gym / Birthdays: " + ", ".join(sorted(extra_taxable)))
    if gateway.get("unsettled_count"):
        notes.append(f"{gateway['unsettled_count']} gateway transaction(s) NOT settled "
                     f"(${gateway['unsettled_amt']:,.2f}) - processed but not completed.")
    if abs(unapplied_gw_net) > 0.001:
        notes.append(f"Unapplied payments/refunds via gateway tenders net to "
                     f"${unapplied_gw_net:,.2f} (included in the Member Fees JE line).")
    if abs(cash_collected) > 0.001:
        notes.append(f"Cash/Check collected this window: ${cash_collected:,.2f}. NOT in this JE - "
                     "deposit it, code the deposit to 2090, and allocate it next month.")
    if abs(external_total) > 0.001:
        notes.append(f"External Credit Card / Nacha payments: ${external_total:,.2f}. NOT in this JE - "
                     "these settle outside the iClassPro gateway (e.g. ClassWallet) and are coded "
                     "straight from the bank feed.")

    return {
        "program_rows": program_rows,
        "ref_taxable": round(ref_taxable, 2),
        "ref_nontaxable": round(ref_nontaxable, 2),
        "fees_section": {
            "Total Collected (Gross)": gateway["gross"],
            "Total Fees": gateway["fees"],
            "Total Net Deposited": gateway["net"],
        },
        "use_tax": round(use_tax, 2),
        "total_tax": total_tax,
        "je_debits": [
            ("2090 Revenue Clearing (iClassPro)", round(gateway["net"], 2)),
            ("6040 Credit Card Merchant Fees", round(gateway["fees"], 2)),
        ],
        "je_credits": [
            ("4200 Sales:Venue (Open Gym + Birthdays)", venue),
            ("4000 Sales:Merchandise (other taxable)", merch),
            ("4100 Sales:Member Fees (non-taxable + unapplied)", member),
            ("2230 Sales Tax Payable", je_tax),
        ],
        "je_balance_ok": abs(je_debits - je_credits) < 0.005,
        "je_debit_total": je_debits,
        "je_credit_total": je_credits,
        "cash_collected": cash_collected,
        "external_total": external_total,
        "notes": notes,
        "taxable_programs": sorted(taxable),
        "nontaxable_programs": sorted(p for p in progs if p not in taxable),
    }


def report_to_grid(report, period_label):
    """Flatten the report into the iClassReport layout (one row per taxable program)."""
    grid = [["iClassReport", period_label, ""],
            ["Program", "Total Collected (Gross)", "Tax Collected (Portion)"]]
    for name, gross, tax in report["program_rows"]:
        grid.append([name, f"{gross:,.2f}", f"{tax:,.2f}"])
    grid.append(["", "", ""])
    grid.append(["Refunds", "", ""])
    grid.append(["Taxable", f"{report['ref_taxable']:,.2f}", ""])
    grid.append(["Non-taxable", f"{report['ref_nontaxable']:,.2f}", ""])
    grid.append(["", "", ""])
    grid.append(["Card Processing Fees", "", ""])
    for k, v in report["fees_section"].items():
        grid.append([k, f"{v:,.2f}", ""])
    grid.append(["", "", ""])
    grid.append(["Use Tax (Gross)", f"{report['use_tax']:,.2f}", ""])
    return grid


# =========================================================================== #
#  UI
# =========================================================================== #
st.title("iClassReport Processor")
st.caption(f"Flip Side Lowell - v{VERSION} - FIN-4 Program Deposit Split + FIN-24 Gateway Transactions to iClassReport")

# ---- STEP 1: basis + which dates to export --------------------------------
st.subheader("Step 1 - Get the reports")
basis = st.radio("Report basis",
                 ["Payout date (recommended)", "Settlement window (transaction date)"],
                 horizontal=True,
                 help="Payout date: iClassPro re-buckets each transaction by when it paid out - "
                      "natively ties to the merchant statement. Settlement window: transaction-date "
                      "exports for a computed window that approximates the same thing.")
PAYOUT_MODE = basis.startswith("Payout")

today = datetime.date.today()
prev_year, prev_month = (today.year - 1, 12) if today.month == 1 else (today.year, today.month - 1)
months = []
y, m = prev_year, prev_month
for _ in range(12):
    months.append((y, m))
    y, m = (y - 1, 12) if m == 1 else (y, m - 1)
month_labels = [datetime.date(y, m, 1).strftime("%B %Y") for y, m in months]
sel = st.selectbox("Month you are closing", month_labels, index=0)
sel_year, sel_month = months[month_labels.index(sel)]
win_start, win_end = settlement_window(sel_year, sel_month)
month_first = datetime.date(sel_year, sel_month, 1)
month_last = (datetime.date(sel_year + 1, 1, 1) if sel_month == 12
              else datetime.date(sel_year, sel_month + 1, 1)) - datetime.timedelta(days=1)

if PAYOUT_MODE:
    st.info(f"**Export the FIN-4 Program Deposit Split for {month_first:%m/%d/%Y} - "
            f"{month_last:%m/%d/%Y} using iClassPro's PAYOUT DATE option.**  \n"
            "It natively matches what the bank received this month. The Gateway report cannot "
            "filter by payout date, so fees come from the merchant statement (Step 3) - or "
            f"optionally upload a transaction-date FIN-24 for {win_start:%m/%d/%Y} - "
            f"{win_end:%m/%d/%Y} to get the same fees with per-transaction detail.")
else:
    st.info(f"**Export both iClassPro reports for:  {win_start:%m/%d/%Y} - {win_end:%m/%d/%Y}**  \n"
            f"(FIN-4 Program Deposit Split and FIN-24 Gateway Transactions, Payments-Received basis. "
            f"This settlement window is what lands in the bank during {sel}, because payouts arrive "
            f"~2 business days after the transaction.)")

# ---- STEP 2: upload ---------------------------------------------------------
st.subheader("Step 2 - Upload")
c1, c2 = st.columns(2)
with c1:
    file_a = st.file_uploader("FIN-4 Program Deposit Split (CSV or XLSX)" if PAYOUT_MODE
                              else "Upload report A (CSV or XLSX)",
                              type=["csv", "xlsx"], key="a")
with c2:
    file_b = st.file_uploader("FIN-24 Gateway Transactions - optional in payout mode" if PAYOUT_MODE
                              else "Upload report B (CSV or XLSX)",
                              type=["csv", "xlsx"], key="b")

with st.expander("Step 3 - Merchant statement totals" +
                 (" (REQUIRED in payout mode if no Gateway file)" if PAYOUT_MODE else " (for exact validation)"),
                 expanded=PAYOUT_MODE):
    st.caption("From the iClassPro monthly merchant statement Summary box.")
    stmt_net_sales = st.number_input("Statement Net Sales (Total Sales - Refunds) $",
                                     min_value=0.0, value=0.0, step=0.01, format="%.2f")
    stmt_fees = st.number_input("Statement Total Processing Fees $ (enter as positive)",
                                min_value=0.0, value=0.0, step=0.01, format="%.2f")
    stmt_net = st.number_input("Statement Net Amount Settled $",
                               min_value=0.0, value=0.0, step=0.01, format="%.2f")
    use_tax = st.number_input("Use Tax (Gross) $", min_value=0.0, value=0.0, step=0.01, format="%.2f")

ready = (file_a is not None and file_b is not None) or \
        (PAYOUT_MODE and (file_a is not None or file_b is not None))

if ready:
    program_rows_raw = gateway_rows_raw = None
    for f in (file_a, file_b):
        if f is None:
            continue
        rows = load_upload(f)
        if rows is None:
            st.stop()
        kind = detect_report_type(rows)
        if kind == "program_split":
            program_rows_raw = rows
        elif kind == "gateway":
            gateway_rows_raw = rows

    if program_rows_raw is None:
        st.error("Need a Program Deposit Split (FIN-4) export - could not identify one "
                 "in the uploaded file(s).")
    elif gateway_rows_raw is None and not PAYOUT_MODE:
        st.error("Settlement-window mode needs the Gateway Transactions (FIN-24) export too. "
                 "Switch to Payout date mode to run without it (fees from the statement).")
    elif gateway_rows_raw is None and PAYOUT_MODE and stmt_fees <= 0:
        st.warning("No Gateway file uploaded - enter the statement's Total Processing Fees in "
                   "Step 3 so the fees section and the 6040 JE line can be built.")
    else:
        progs, unapplied_gw_net, period = parse_program_split(program_rows_raw)
        fin4_total = round(sum(v["total"] for v in progs.values()) + unapplied_gw_net, 2)

        if gateway_rows_raw is not None:
            gateway = parse_gateway(gateway_rows_raw)
        else:
            # Payout mode without FIN-24: statement supplies the fees.
            gateway = {
                "gross": fin4_total,
                "fees": round(stmt_fees, 2),
                "net": round(fin4_total - stmt_fees, 2),
                "count": 0, "settled_net": round(fin4_total - stmt_fees, 2),
                "unsettled_amt": 0.0, "unsettled_count": 0,
                "period": period, "date_min": None, "date_max": None,
            }
        if PAYOUT_MODE and gateway_rows_raw is not None:
            # FIN-24 is transaction-date. Only a settlement-window export carries the
            # right fees; a calendar-month export does not. Statement fees always win.
            gateway = dict(gateway)
            gateway["gross"] = fin4_total
            if stmt_fees > 0 and abs(gateway["fees"] - stmt_fees) > 0.005:
                st.warning(f"The uploaded FIN-24 shows fees of ${gateway['fees']:,.2f}, but the "
                           f"statement says ${stmt_fees:,.2f} - the FIN-24 is probably a "
                           f"calendar-month export (transaction basis), which covers a different "
                           f"set of payouts than this month's statement. **Using the statement's "
                           f"fees.** For matching per-transaction fee detail, export FIN-24 for "
                           f"{win_start:%m/%d/%Y} - {win_end:%m/%d/%Y} instead.")
                gateway["fees"] = round(stmt_fees, 2)
            elif stmt_fees <= 0 and gateway.get("date_min") and gateway["date_min"].day == 1:
                st.warning("The FIN-24 looks like a calendar-month export - its fees are on "
                           "transaction basis and will NOT match the statement. Enter the "
                           "statement's Total Processing Fees in Step 3 to correct this.")
            gateway["net"] = round(fin4_total - gateway["fees"], 2)

        report = build_iclassreport(progs, unapplied_gw_net, gateway, use_tax=use_tax)
        period_label = period or gateway["period"] or datetime.date.today().strftime("%m/%d/%Y")

        src_desc = "FIN-4 payout basis" if PAYOUT_MODE else "settlement window"
        st.success(f"Processed period: {period_label} ({src_desc})  -  {len(progs)} programs" +
                   (f"  -  {gateway['count']} gateway transactions" if gateway.get("count") else ""))

        # ---- window sanity check (settlement mode only) ----
        if not PAYOUT_MODE and gateway.get("date_min") and gateway.get("date_max"):
            if gateway["date_min"] != win_start or gateway["date_max"] != win_end:
                if gateway["date_min"].day == 1:
                    st.warning(f"The uploaded reports cover **{gateway['date_min']:%m/%d/%Y} - "
                               f"{gateway['date_max']:%m/%d/%Y}**, which looks like a calendar month, "
                               f"not the settlement window **{win_start:%m/%d/%Y} - {win_end:%m/%d/%Y}**. "
                               "Re-export with the Step 1 dates, or switch to Payout date mode.")
                else:
                    st.warning(f"Uploaded report dates ({gateway['date_min']:%m/%d/%Y} - "
                               f"{gateway['date_max']:%m/%d/%Y}) don't exactly match the Step 1 window "
                               f"({win_start:%m/%d/%Y} - {win_end:%m/%d/%Y}). Double-check the month "
                               "selected and the export range.")

        # ---- iClassReport table ----
        st.subheader("iClassReport")
        grid = report_to_grid(report, period_label)
        df_display = pd.DataFrame(grid[2:], columns=grid[1])
        st.dataframe(df_display, width="stretch", hide_index=True)

        # ---- statement validation ----
        st.subheader("Merchant statement validation")
        if stmt_net_sales > 0 or stmt_fees > 0 or stmt_net > 0:
            checks = [
                ("Net Sales", gateway["gross"], stmt_net_sales),
                ("Processing Fees", gateway["fees"], stmt_fees),
                ("Net Settled", gateway["net"], stmt_net),
            ]
            all_ok = True
            for label, got, want in checks:
                if want <= 0:
                    continue
                var = round(got - want, 2)
                if abs(var) < 0.005:
                    st.success(f"{label}: report ${got:,.2f} = statement ${want:,.2f}  (exact match)")
                else:
                    all_ok = False
                    st.error(f"{label}: report ${got:,.2f} vs statement ${want:,.2f}  ->  "
                             f"variance ${var:,.2f}. This should be $0.00. Likely causes: wrong "
                             "export range/basis, or (settlement mode) a transaction that slipped "
                             "a batch at the window edge.")
            if all_ok:
                st.success("All statement totals match exactly. Report is good to send and the JE "
                           "is good to post.")
            st.caption("Note: the statement's separate 'Bank Transfer Fees' line (e.g. $0.54) is an "
                       "account fee, not a processing fee - it is excluded from this check and does "
                       "not reduce the payouts.")
        else:
            st.info("Enter the merchant statement totals in Step 3 to validate. These should match "
                    "to the penny.")

        # ---- draft JE ----
        st.subheader("Draft 2090 clearing JE (gateway-settled money only)")
        st.caption("Post in Xero dated the last day of the month. Cash and External CC/Nacha are "
                   "intentionally excluded - see notes.")
        je_rows = [(k, f"{v:,.2f}", "") for k, v in report["je_debits"]] + \
                  [(k, "", f"{v:,.2f}") for k, v in report["je_credits"]] + \
                  [("TOTALS", f"{report['je_debit_total']:,.2f}", f"{report['je_credit_total']:,.2f}")]
        st.table(pd.DataFrame(je_rows, columns=["Account", "Debit", "Credit"]))
        if report["je_balance_ok"]:
            st.success("JE balances: debits = credits.")
        else:
            st.error(f"JE DOES NOT BALANCE (DR ${report['je_debit_total']:,.2f} vs "
                     f"CR ${report['je_credit_total']:,.2f}). Do not post - check the export files "
                     "and Step 3 fee entry.")

        # ---- notes ----
        if report["notes"]:
            st.subheader("Notes / exceptions")
            for note in report["notes"]:
                st.write("- " + note)
        if PAYOUT_MODE:
            st.write("- Payout basis: cash sales never appear on this report (cash has no payout). "
                     "Capture the cash split when you make the monthly cash deposit.")

        # ---- export: copy/paste + download ----
        st.subheader("Copy & paste")
        st.caption("Hover the box and click the copy icon (top-right), then paste into Google Sheets "
                   "or Excel - it splits into columns automatically.")
        tsv = "\n".join("\t".join(str(c) for c in row) for row in grid)
        st.code(tsv, language=None)

        st.caption("Draft 2090 clearing JE:")
        je_tsv = "\n".join([f"{k}\t{v:,.2f}\t" for k, v in report["je_debits"]] +
                           [f"{k}\t\t{v:,.2f}" for k, v in report["je_credits"]])
        st.code(je_tsv, language=None)

        csv_buf = io.StringIO()
        csv.writer(csv_buf).writerows(grid)
        st.download_button("Download iClassReport (CSV)", csv_buf.getvalue(),
                           file_name=f"iClassReport_{period_label.replace('/', '-')}.csv",
                           mime="text/csv")
else:
    st.info("Upload the export(s) to begin. Order doesn't matter - the app detects which is which.")
